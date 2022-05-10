import pandas as pd
import sys
import openpyxl as oxl
sys.path.append(r"C:\Users\Usuario\Documents\Trabajo\TU Berlin\E-mobility flexiblity potential\mob4anci")
from m4a_classes.inputclass import inputInfo




def readInput(path):

    data = pd.read_excel(path, 'input')
    data = data.drop(data[data.readinput==0].index)
    
    dayoption = ['week', 'sat', 'hol'] #workweek, saturday, holiday/sunday

    aktInput = [None]*data.shape[0] #initialize empty list. I dont initialize with the 
    #class already cause python will just create a copy of the class and not 2 independent
    
    for j in range(len(aktInput)):
        
        aktInput[j] = inputInfo()
        items = list(aktInput[j].__dict__.keys())

        for element in items:

            if element =='connectivity':

                cols = [aktInput[j].id + '_' + day for day in dayoption]
                connectivity = pd.read_excel(path, 'connectivity', index_col=0).loc[:,cols]

            else:

                setattr(aktInput[j], element, data[element].iloc[j])

        aktInput[j].connectivity = connectivity



    return aktInput

def readAnService(path, servicetype):

    data = pd.read_excel(path, servicetype)

    return data

def writeOutput(path, id, servicetype, checks):
    ''' get the path and opern the workbook. Then, we get the checks that we have stored in the checks list obtaine from comaperValues.
    '''
    wb = oxl.load_workbook(path)
    wsname = wb.sheetnames[0]
    ws = wb[wsname]
    
    for j in range(len(checks[0])):
        
        rowindex = ws.max_row + 1

        for colindex in range(1,ws.max_column+1):

            #first column is always id

            if colindex == 1:

                ws.cell(row=rowindex, column=colindex).value = id

            #second column is always service type

            elif colindex == 2:

                ws.cell(row=rowindex,column=colindex).value = servicetype

            #add the product and the boolean checks

            else:
                # check if it is string or not, and write
                
                if not isinstance(checks[colindex-3][j], str):
                    
                    ws.cell(row=rowindex,column=colindex).value = str(checks[colindex-3][j])
                    
                else:
                    
                    ws.cell(row=rowindex,column=colindex).value = checks[colindex-3][j]
                

        
    wb.save(path)





