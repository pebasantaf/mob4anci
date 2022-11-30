import os
import m4a_checker.checkfunctions as prc
#from m4a_classes.models import VirtualPowerPlant
from pathlib import Path
import pandas as pd
import openpyxl as oxl


def updateListofLists(list,value,index,pos):

    # one index for each freq_control product and one pos for each check
    if index == 0:

        list[pos] = [value]

    elif index > 0:

        list[pos] +=[value]

    return list

def writeOutput(path, id, servicetype, checks):

    wb = oxl.load_workbook(path)
    wsname = wb.sheetnames[0]
    ws = wb[wsname]
    
    for j in range(len(checks[0])):
        
        rowindex = ws.max_row + 1

        for colindex in range(1,ws.max_column+1):

            #first column is always id

            if colindex == 1:

                ws.cell(row=rowindex, column=colindex).value = id

            #second column is always service type

            elif colindex == 2:

                ws.cell(row=rowindex,column=colindex).value = servicetype

            #add the product and the boolean checks

            else:
                # check if it is string or not, and write
                
                if not isinstance(checks[colindex-3][j], str):
                    
                    ws.cell(row=rowindex,column=colindex).value = str(checks[colindex-3][j])
                    
                else:
                    
                    ws.cell(row=rowindex,column=colindex).value = checks[colindex-3][j]
                

        
    wb.save(path)


def getLPFileDirectories(file_directionary: str,filename: str) -> dict:
    ## set objective (Minimize/Maximize)
    
    obj_minmax = 'Minimize'
    
    
    ## get file directories
    
    # main file
    lp= file_directionary + '\\'+'LPfiles\\'+ filename+'.lp'
    
    # temporary file for objective variables
    filename_obj = filename +'_obj'
    obj = file_directionary + '\\'+'LPfiles\\'+ filename_obj+'.lp'
    with open(obj,'w+') as f:
        f.truncate(0)
        f.write(obj_minmax +'\n')

    
    # temporary file for constraints
    filename_cons = filename +'_cons'
    cons = file_directionary + '\\'+'LPfiles\\'+ filename_cons+'.lp'
    with open(cons,'w+') as f:
        f.truncate(0)
        f.write('\nSubject To\n')
    
    # temporary file for boundaries
    filename_bounds = filename +'_bounds'
    bounds = file_directionary + '\\'+'LPfiles\\'+ filename_bounds+'.lp'
    with open(bounds,'w+') as f:
        f.truncate(0)
        f.write( '\nBounds\n')
    
    # temporary file for binaries
    filename_binaries = filename +'_binaries'
    binaries =file_directionary + '\\'+'LPfiles\\'+ filename_binaries+'.lp'
    with open(binaries,'w+') as f:
        f.truncate(0)
        f.write('\nBinaries\n')

    return {'lp':lp,'obj':obj,'cons':cons,'bounds':bounds,'binaries':binaries}

# -*- coding: utf-8 -*-
"""
Info
----
This file contains the basic functionalities of the MergeFiles
class.
"""
def mergeFiles(filedirectory):
    obj = ''
    cons =''
    binaries =''
    bounds = ''
    with open(filedirectory['obj'],'r') as f:
        obj =f.read()
        f.close()
    with open(filedirectory['cons'],'r') as f:
        cons =f.read()
        f.close()
    with open(filedirectory['binaries'],'r') as f:
        binaries =f.read()
        f.close()
    with open(filedirectory['bounds'],'r') as f:
        bounds =f.read()
        f.close()
    with open(filedirectory['lp'],'a+') as f:
        f.truncate(0)
        f.write(obj)
        f.write(cons)
        f.write(bounds)
        f.write(binaries)
        
        f.write('\nEnd')
        f.close()
        