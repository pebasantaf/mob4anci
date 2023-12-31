import statistics
import numpy as np
import pandas as pd
import os
from pathlib import Path
import m4a_checker.checkfunctions as pcr
from utils.utils import updateListofLists, writeOutput, getLPFileDirectories, mergeFiles
import cplex
from m4a_optimizer.writeLPfiles import *
from m4a_optimizer.setCS import setCS
from openpyxl import workbook, load_workbook
import scipy.io as scio
from datetime import datetime

class VirtualPowerPlant:
    
    #filedirectory = os.path.join(os.getcwd(), "data\LPfiles")
    
    def cplexOptimization(self, directories):
        
        # initialize CPLEX object
        cpx = cplex.Cplex()
        
        # read optimization model
        cpx.read(directories['lp'])
        
        # deactivate creation of log-file
        cpx.parameters.output.clonelog.Cur = 4
        
        # choose LP method
        # 0 - Automatic: Let Cplex choose
        # 1 - Primal Simplex
        # 2 - Dual Simplex
        # 3 - Network Simplex
        # 4 - Barrier
        # 5 - Sifting
        # 6 - Concurrent (Dual, Barrier and Primal in opportunistic parallel mode)
        cpx.parameters.lpmethod.Cur = 0
        
        # solve optimization problem
        cpx.solve()
        
        # get variable names and respective solution vector
        """if (isfield(cpx.Solution,'x')):
            cpx_sol = cpx.Solution.x
            cpx_var = cpx.Model.colname
        else:
            error(sprintf(['\nNo optimal soluation available, please adjust optimization model and check command window output for more info!']))
        """
        return cpx
    
class FleetOperator:
    
    path = os.path.join(Path(os.getcwd()).parent, 'inputvalues.xlsx')
    outpath = os.path.join(Path(os.getcwd()).parent, 'outputvalues.xlsx')
    lpfiledirectory = os.path.join(os.getcwd(), "data")
    
    def demandmat2excel(self, matpath, inputid):
        
        matfile = scio.loadmat(matpath)
        
        valuesresolution = matfile['iLMS2VPP_week'][0,0] #<-- set the array you want to access. 
        keys = matfile['iLMS2VPP_week'][0,0].dtype.descr

        timeseries = valuesresolution['tincr_60'][0][0]
        matfilecolumns = valuesresolution['tincr_60'][0].dtype.descr
        columns = ['E_d_esp', 'E_d_lep', 'E_d_flex', 'P_d_max', 'P_d_nonopt']
        cols2write = [col + '_' + inputid for col in columns]
        
        wb = load_workbook(self.path)
        sheet = wb['demand']
        
        for i in range(len(columns)):
            
            maxcol = sheet.max_column
            
            sheet.cell(row=1, column=maxcol+1).value = cols2write[i]
            
            
            listofvalues = list(np.concatenate(np.abs(timeseries[i]), axis=0))
            
            for j in range(len(listofvalues)):
                
                sheet.cell(row=j+2, column=maxcol+1).value = listofvalues[j]
        
        wb.save(self.path)

        '''
        # Assemble the keys and values into variables with the same name as that used in MATLAB
        for i in range(len(keys)):
            key = keys[i][0]
            val = np.squeeze(vals[key][0][0])  # squeeze is used to covert matlat (1,n) arrays into numpy (1,) arrays. 
            exec(key + '=val')
        print()
        
        '''
    def createFleetObjects(self): 
        
        #read input values for the fleet and demand values
        data = pd.read_excel(self.path, 'input')
        data = data.drop(data[data.readinput==0].index)
        
        demand =  pd.read_excel(self.path, 'demand')
        
        #get number of timesteps to limit size of optimization
        
        CSset =setCS()
        limittime = CSset.nr_timesteps
        
        dayoption = ['week', 'sat', 'hol'] #workweek, saturday, holiday/sunday
        
        fleetobjects = dict.fromkeys(data.id.values)
        
        for key in fleetobjects:

            fleetobjects[key] = electricFleet() #store the object in the dictionary
            items = fleetobjects[key].__dict__.keys() #get all the attributes of the object


            for variable in items:

                if variable =='connectivity': #if we wanna get the connectivity, we use a special call cause it is a larger dataset

                    cols = [key + '_' + day for day in dayoption]
                    connectivity = pd.read_excel(self.path, 'connectivity', index_col=0).loc[:,cols]

                elif 'E_d' in variable or 'P_d' in variable:
                    
                    #get list of columns with correct id
                    
                    idcolumns = [col for col in demand.columns if 'bus1' in col]
                    currentdf = demand[idcolumns]

                    currentvariable = [v for v in idcolumns if variable in v]
                    
                    valueslistoflist = currentdf[currentvariable].values.tolist()[:limittime]
                    
                    goodvalues = [item for sublist in valueslistoflist for item in sublist]
                    
                    setattr(fleetobjects[key], variable ,goodvalues)

                    
                
                else: #for the rest of the elements, just set the value
                    
                    index = data.index[data['id']==key][0] #get the index position of the corresponding id
                    setattr(fleetobjects[key], variable, data[variable].iloc[index]) #set the attribute in the fleet object

            fleetobjects[key].connectivity = connectivity #apply connectivity

        return fleetobjects
    
    
    
    def compareValues(self, inputdata, servicedata, servicetype, writecols):

        checks = [['-']]*len(writecols)

        # updating inputdata considering negative/positive power balancing

        if servicetype == 'freq_control':
            
            # 4 hours each participation window for frequency control
            windowtime = 4
            index = 0
            limres = False

            # for every of the three frequency reserve products

            for key in servicedata:
                
                realpower = inputdata.getRealValue('power',servicedata[key].reservetype,True)
                
                # get the real energy
                realenergy = inputdata.getRealValue('energy', servicedata[key].reservetype, True)
                
                # check whether mk power = pq power

                mkpower = realenergy/windowtime 

                # checking whether limited storage requirements apply

                if mkpower < inputdata.getRealValue('power', servicedata[key].reservetype, False): #mk power smaller than the real power without considering connectivity

                    print("\nBalancing reserve provider with limited storage capacity. Marketable power of: " + str(mkpower) + "MW")
                    limres = True
                
                
                print("\nAnalyzing " + servicedata[key].product + " product.")
                
                # add product to checks list
                
                checks = updateListofLists(checks, servicedata[key].product, index, writecols.index('product'))
                checks = updateListofLists(checks, servicedata[key].reservetype, index, writecols.index('reservetype'))
                
                #if the input its a dash, store weeksection and timeslice/hour of the max. connectivity because it is the one that is going to be used
                
                if inputdata.weeksection == '-' or inputdata.timeslice == '-': 
                    
                    checks = updateListofLists(checks, inputdata.getMaxWindow()[1][1], index, writecols.index('weeksection'))
                    checks = updateListofLists(checks, inputdata.getMaxWindow()[1][0], index, writecols.index('timeslice'))

                else: # else, you have input a specific timeslice/hour. Therefore, a dash is given as an ouput as that is the value already being used and not a certain maximum is necessary
                    
                    checks = updateListofLists(checks, '-', index, writecols.index('weeksection'))
                    checks = updateListofLists(checks, '-', index, writecols.index('timeslice'))
                    
                
                # CHECK1: check minimum bidding power is fulfilled
                
                checks = pcr.check1freqcontrol(realpower, servicedata[key].minimum_bid_power, index, checks, writecols)

                
                # CHECK2: check if ramping requirement is fulfilled
                
                checks = pcr.check2freqcontrol(inputdata.rampup, servicedata[key].ramp_up, index, checks, writecols)


                # check if there is the reservoir is limited

                if limres:
                    
                    # CHECK3: minimum energy to PQ and MK energy capacity
                    
                    checks = pcr.check3freqcontrol(realenergy, realpower, servicedata[key], index, mkpower, checks, writecols)
                    
                    # CHECK4: Maximum power when limited reservoir
                    
                    checks = pcr.check4freqcontrol(realpower, index, mkpower, checks, writecols)
                    
                index += 1    
             # Check redispatch power
    
        elif servicetype == 'redispatch':
        
            checks = pcr.checkredispatchpower(checks, inputdata, servicedata, writecols)
        
        
        writeOutput(self.outpath, inputdata.id, servicetype, checks)
        
        return checks

    def setValuesForOptimization(self):   # NOTE: UNIFINSHED
        
        #Vpp_name_list =['fleet','em']
        results = []
        em = electricityMarket()
        em.setGeneralMarketAttributes()
        em.setBalancingMarketAttributes()
        
        fleet = self.createFleetObjects()
        
        directories = getLPFileDirectories(self.lpfiledirectory,'VPP')
        cs = setCS()
        
        writeLPem(directories, cs, em)
        writeLPbm(directories, cs, em)
        writeLPadd(directories, cs)
        
        for key in fleet:
            
            writeLPfleet(directories, cs, fleet[key])
            
            mergeFiles(directories)
            
            vpp = VirtualPowerPlant()
            
            results += [vpp.cplexOptimization(directories)]
        
        return results, fleet, em


    
class electricFleet:
    
    windowtime = 4
    
    def __init__(self):
        self.id = None
        self.storagetype = None
        self.chargetype = None
        self.chargeefficiency = None
        self.dischargeefficiency = None
        self.energy = None
        self.power = None
        self.remotecontrol = None
        self.rampup = None
        self.weeksection = None
        self.timeslice = None
        self.connectivity = None
        self.E_d_esp = None
        self.E_d_lep = None
        self.E_d_flex = None
        self.P_d_max = None
        self.P_d_nonopt = None
        self.cost = None
        self.ghg_CO2 = None
        self.task = None
    
        
    def getRealValue(self, magnitude, flow, connectivity):
        
        #energy or power
        if magnitude == 'energy':
            
            value = self.energy
            
        elif magnitude == 'power':
            
            value = self.power
        
        #positive, negative or symmetric
        if flow == 'positive':
            
            realvalue = round(value * self.dischargeefficiency, 4)
            
        elif flow == 'negative':
            
            realvalue = round(value * self.chargeefficiency, 4)
            
        elif flow == 'symmetric':

            realvalue = round(value * statistics.mean([self.chargeefficiency, self.dischargeefficiency]), 4)
            
        else:
            
            raise NameError("Incorrect input string for checkControlReserve")
        
        # if we want also the multiplication times connectivity, set True
        if connectivity:
            
            # if we dont select time slice or weeksection, get the best value
            
            if self.weeksection == '-' or self.timeslice == '-':
                
                realvalue = realvalue * self.getMaxWindow()[0]
                
            else:
                
                realvalue = realvalue * self.getMinConnectivityinWindow()
        
        return realvalue
    
    def getMaxConnectivity(self):
        
        #get the maxium connectivity value of all value
        maxcon = max(self.connectivity.max())
        maxcon_loc = [self.connectivity.where(self.connectivity==maxcon).dropna().index[0], 
                      self.connectivity.where(self.connectivity==maxcon).dropna().columns[0].split('_')[1]]
        
        return maxcon, maxcon_loc
    
    def getMinConnectivityinWindow(self):
        
        #get the minimum value of a 4 hour window of connectivity
        currentwindow = self.connectivity[self.id + "_" + self.weeksection][(self.timeslice-1)
        *self.windowtime:self.timeslice*self.windowtime]
        minconnect = min(currentwindow)
        
        return minconnect

        
    def getMaxWindow(self): #check all the slots, for each slot find the minimum value of a certain slot and get the value of the slot with the highest minimum value
        
        #initialize maximum value and calculate slots (6 slots)
        fullmax = 0
        slots = int(self.connectivity.shape[0]/self.windowtime)
        
        #for each slot
        for i in range(slots):
            
            # we get 6 4x3 matrices. For each slot, we want to know the minimum for each column (1x3 vector) BUT we want the maximum of that
            # This way, we initially get the limiting connectivity which will determine the energy, but we get the highest minimum
            aktmax = max(self.connectivity[i*self.windowtime:(i+1)*self.windowtime].min(axis=0))
            
            #get the index, it being [slot, weeksection]
            aktindexmax = [i, self.connectivity.columns[np.argmax(self.connectivity[i*self.windowtime:(i+1)*self.windowtime].min(axis=0))].split('_')[1]]

            
            
            # update if we find a higher value
            if aktmax > fullmax:
                
                fullmax = aktmax
                fullindexmax = aktindexmax
        
        
             
        return fullmax, fullindexmax
        


        
class electricityMarket:
    
    file = pd.ExcelFile(os.path.join(Path(os.getcwd()).parent, 'electricitymarketvalues.xlsx'))
    
    def __init__(
        self,
        P_im_min = None,
        P_im_max = None,
        P_ex_min = None,
        P_ex_max = None,
        ghg_CO2_IM = None,
        ghg_CO2_EX = None,
        price_em = None,
        price_FCR_cap=None,
        price_aFRR_cap_pos= None,
        price_aFRR_cap_neg= None,
        price_mFRR_cap = None,
        price_aFRR_en_pos = None,
        price_aFRR_en_neg = None,
        price_mFRR_en = None
        
    ):
        
        # Call to super class
        #super(emAttributes, self).__init__(
         #   )
        """
        Info
        ----
        This class provides a model with the basic attributes of a
        energy market.
        """
        # min. power import (kW)
        self.P_im_min = 0;
        
        # max. power import (kW)
        self.P_im_max = 1000000;
        
        # min. power export (kW)
        self.P_ex_min = 0;
        
        # max. power export (kW)
        self.P_ex_max = 1000000;
        
        # greenhouse gas emissions for market import (gCO2/kWh)
        self.ghg_CO2_IM = 559;
        
        # greenhouse gas emissions for market export (gCO2/kWh)
        self.ghg_CO2_EX = 0.001;

        self.price_em =  None
        
        self.price_FCR_cap = None
        self.price_aFRR_cap_pos = None
        self.price_aFRR_cap_neg = None
        self.price_mFRR_cap = None
        self.price_aFRR_en_pos = None
        self.price_aFRR_en_neg = None
        self.price_mFRR_en = None
        
    def setGeneralMarketAttributes(self):
        
        generaldata = pd.read_excel(self.file, 'general_data')
        
        for key in generaldata.columns:
            
            setattr(self, key, generaldata.loc[0,key])

            
        #set market price
        
        CSset =setCS()
        limittime = CSset.nr_timesteps
        price = pd.read_excel(self.file, 'spot_market')
        

        self.price_em = price.spot_price.values.tolist()[:limittime]
     
    
    def storeReserveMarketData(self, datapaths):
        
        #datapath must be a list of paths
        
        targetdf = pd.read_excel(self.file, 'balancing_market')
        
        excel_workbook = load_workbook(self.file)
        
        with pd.ExcelWriter(self.file, engine='openpyxl', mode="a", if_sheet_exists='overlay') as writer:
            
            writer.book = excel_workbook
            #writer.sheets = dict((ws.title, ws) for ws in excel_workbook.worksheets)
        
            for path in datapaths:
                
                data = pd.read_csv(path, sep=';')
                
                data = data.replace('-',0, regex=True) #replace - with zeros. For us , if there was no volume traded it is the same as a 0
                
                # turn date and time into datetime format
                
                data['Date'] = pd.to_datetime(data['Date'])
                
                data['Time of day'] = pd.to_datetime(data['Time of day']).dt.time
                
                # turn every object into numeric
                
                for col in data.columns[2:]:
                    
                    datatype = data[col].dtype
                    
                    if datatype == 'O':
                        
                        try:
                        
                            data[col] = pd.to_numeric(data[col])
                        
                        except ValueError:
                            
                            data[col] = data[col].str.replace(',', '').astype(float)
                
                # fill nans with zeros
                
                data = data.fillna(0)
                        
                #NOTE: Add data 4 by 4 to get hourly. then store. Important: Values change so much every 15 min cause volume if bidded for 4 hours but might be activated or not. 

                hourlydata = pd.DataFrame(columns=data.columns)
                
                timestamps =  pd.date_range(datetime.combine(data["Date"][0].date(), data["Time of day"][0]), periods=7*24, freq="60min")
                
                hourlydata["Date"] = timestamps.date
                hourlydata["Time of day"] = timestamps.time
                
                
                #select columns: columns with power or energy, we add. columns with price, we make an average
                
                powermask = ['Volume' in col for col in data.columns]
                powercols = data.columns[powermask]
                
                pricemask = ['price' in col for col in data.columns]
                pricecols = data.columns[pricemask]
                
                for index,row in hourlydata.iterrows():
            
                    row[powercols] = sum(data.loc[index*4:(index+1)*4-1,powercols].values)
                    row[pricecols] =  np.mean(data.loc[(index)*4:(index+1)*4-1,pricecols].values, axis=0)
                    
                # different cases for FCR, aFRR and mFRR     
                
                if 'Automatic_Frequency_Restoration_Reserve' in path:
                    
                    if len(targetdf.date) == 0 and len(targetdf.time) == 0:
                        
                        targetdf.date = hourlydata["Date"]

                        targetdf.time = hourlydata["Time of day"]
                        
                    
                    for col in hourlydata.columns[2:]:
                        
                        targetdf['aFRR-' + col] = hourlydata[col]


            targetdf.to_excel(writer, 'balancing_market', index=False)
            
            #writer.save() #Adding this line produces a corruppted excel that must be repaird
                    
            return targetdf
            
            
    def setBalancingMarketAttributes(self):
        
        bmdata = pd.read_excel(self.file, 'balancing_market')
        
        colmask = ['aFRR' in col for col in bmdata.columns]
        
        #this is momentaneously hardcoded
        
        self.price_aFRR_cap_pos = bmdata['aFRR-Procurement price (+)[€/MW]'].values
        self.price_aFRR_cap_neg = bmdata['aFRR-Procurement price (-)[€/MW]'].values
        
        self.price_aFRR_en_pos = bmdata['aFRR-Activation price (+)[€/MWh]'].values
        self.price_aFRR_en_neg = bmdata['aFRR-Activation price (-)[€/MWh]'].values
        
        print()
        
        
    def setReserveMarketAttributes(self):  # NOTE: UNIFINSHED
        

        reserveprice = pd.read_excel(self.file, 'balancing_market')
        
        print()
            


        
class ancillaryService:
    
    path = os.path.join(Path(os.getcwd()).parent, 'ancillaryservicevalues.xlsx')

        
    def getFreqcontrolObject(self):
        
        #get data for frquency control
        data = pd.read_excel(self.path, 'freq_control')
        
        #get names for initializing dictionary
        a = data["product"].values.tolist()
        b = [s[0:3] for s in data.reservetype.values.tolist()]
        dictnames = ["{}_{}".format(a_, b_) for a_, b_ in zip(a, b)]
        
        productdict = dict.fromkeys(dictnames, None)

        #create an object for each product and store its attributes
        
        for key in productdict:
            
            productdict[key] = freqControl()
            items = list(productdict[key].__dict__.keys())
            
            for element in items:
                
                index = dictnames.index(key)
                setattr(productdict[key], element, data[element].iloc[index])
                

        return productdict
    
     
    def getRedispatchObject(self):
        
        #get data for frquency control
        data = pd.read_excel(self.path, 'redispatch')

        product = redispatch()
        product.minimum_bid_power_1 = data[list(product.__dict__.keys())[0]][0]
        product.minimum_bid_power_2 = data[list(product.__dict__.keys())[1]][0]
        productdict = dict.fromkeys(data["product"], product)
        
        return productdict
        

        
class freqControl(ancillaryService):
    
    def __init__(self):
        
        self.product = None
        self.reservetype = None
        self.max_load_time = None
        self.minimum_bid_power = None
        self.ramp_up = None
        self.min_cap_PQ=None
        self.min_cap_MK_1=None
        self.min_cap_MK_2=None
        self.min_cap_MK_3=None
        
        
class redispatch(ancillaryService):
    
    def __init__(self):
        
        self.minimum_bid_power_1 = None
        self.minimum_bid_power_2 = None

        
    
        
        
        

        
        